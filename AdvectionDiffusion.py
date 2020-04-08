import numpy as np
import openpyxl
import math 
import random

#---------------Data Generation-----------------
# u_t == -Vx * u_x -Vy * u_y -Vz * u_z + D * (u_xx + u_yy + u_zz)
# velocity
Vx = 1
Vy = 2
Vz = 1

# diffusivities
D = 1

# noise
noise = 20

# numerical solution
time = []
u_t_noise = []
u_noise = []
u_x_noise = []
u_y_noise = []
u_z_noise = []
u_xx_noise = []
u_yy_noise = []
u_zz_noise = []

# main
# noise 0 workbook
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.cell(1,1).value = 'x'
ws1.cell(1,2).value = 'y'
ws1.cell(1,3).value = 'z'
ws1.cell(1,4).value = 't'
ws1.cell(1,5).value = 'u'
ws1.cell(1,6).value = 'u_t'
ws1.cell(1,7).value = 'u_x'
ws1.cell(1,8).value = 'u_y'
ws1.cell(1,9).value = 'u_z'
ws1.cell(1,10).value = 'u_xx'
ws1.cell(1,11).value = 'u_yy'
ws1.cell(1,12).value = 'u_zz'
# noise 10 workbook
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.cell(1,1).value = 'x'
ws2.cell(1,2).value = 'y'
ws2.cell(1,3).value = 'z'
ws2.cell(1,4).value = 't'
ws2.cell(1,5).value = 'u'
ws2.cell(1,6).value = 'u_t'
ws2.cell(1,7).value = 'u_x'
ws2.cell(1,8).value = 'u_y'
ws2.cell(1,9).value = 'u_z'
ws2.cell(1,10).value = 'u_xx'
ws2.cell(1,11).value = 'u_yy'
ws2.cell(1,12).value = 'u_zz'
# change for 100 different initial conditions --> center (x0, y0, z0)
for i in range(0,100,1):
    x0, y0, z0 = random.random(), random.random(), random.random()
    # numbers of dataset in each initial conditions --> position (x, y, z)
    for j in range(0,5,1):
        x, y, z = random.random(), random.random(), random.random()
        # the noise is 0 or 10%
        for n in range(0,noise,10):
            # take 10 time steps on each position
            for t in range(0,10,1):
                # exact solution in (x, y, z, t) referring to initial condition (x0, y0, z0)
                u = 1/((4*t+1)**1.5)*math.exp((-(x-Vx*t-0.5)**2/(D*(4*t+1))-(y-Vy*t-0.5)**2/(D*(4*t+1))-(z-Vz*t-0.5)**2/(D*(4*t+1))))
                u_t = (math.sqrt(4*t+1)*(4*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D))*z**2+(2*Vz-4)*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D))*z+(4*math.exp(((2*Vx*t+1)*x)/(4*D*t+D))*y**2+(2*Vy-4)*math.exp(((2*Vx*t+1)*x)/(4*D*t+D))*y+(4*x**2+(2*Vx-4)*x+((-4*Vz**2)-4*Vy**2-4*Vx**2)*t**2+((-2*Vz**2)-2*Vy**2-2*Vx**2-24*D)*t-Vz-Vy-Vx-6*D+3)*math.exp(((2*Vx*t+1)*x)/(4*D*t+D)))*math.exp(((2*Vy*t+1)*y)/(4*D*t+D)))*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+4*x**2+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(256*D*t**4+256*D*t**3+96*D*t**2+16*D*t+D)
                u_x = -((2*x-2*Vx*t-1)*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+((-8*Vy*t)-4)*y+4*x**2+((-8*Vx*t)-4)*x+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(math.sqrt(4*t+1)*(16*D*t**2+8*D*t+D))
                u_y = -((2*math.exp(((2*Vx*t+1)*x)/(4*D*t+D))*y+((-2*Vy*t)-1)*math.exp(((2*Vx*t+1)*x)/(4*D*t+D)))*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+((-8*Vy*t)-4)*y+4*x**2+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(math.sqrt(4*t+1)*(16*D*t**2+8*D*t+D))
                u_z = -((2*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D))*z+((-2*Vz*t)-1)*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D)))*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+4*x**2+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(math.sqrt(4*t+1)*(16*D*t**2+8*D*t+D))
                u_xx = (math.sqrt(4*t+1)*(4*x**2+((-8*Vx*t)-4)*x+4*Vx**2*t**2+(4*Vx-8*D)*t-2*D+1)*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+((-8*Vy*t)-4)*y+4*x**2+((-8*Vx*t)-4)*x+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(256*D**2*t**4+256*D**2*t**3+96*D**2*t**2+16*D**2*t+D**2)
                u_yy = (math.sqrt(4*t+1)*(4*math.exp(((2*Vx*t+1)*x)/(4*D*t+D))*y**2+((-8*Vy*t)-4)*math.exp(((2*Vx*t+1)*x)/(4*D*t+D))*y+(4*Vy**2*t**2+(4*Vy-8*D)*t-2*D+1)*math.exp(((2*Vx*t+1)*x)/(4*D*t+D)))*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+((-8*Vy*t)-4)*y+4*x**2+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(256*D**2*t**4+256*D**2*t**3+96*D**2*t**2+16*D**2*t+D**2)
                u_zz = (math.sqrt(4*t+1)*(4*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D))*z**2+((-8*Vz*t)-4)*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D))*z+(4*Vz**2*t**2+(4*Vz-8*D)*t-2*D+1)*math.exp(((2*Vy*t+1)*y+(2*Vx*t+1)*x)/(4*D*t+D)))*math.exp(-((4*z**2+((-8*Vz*t)-4)*z+4*y**2+4*x**2+(4*Vz**2+4*Vy**2+4*Vx**2)*t**2+(4*Vz+4*Vy+4*Vx)*t+3)/(16*D*t+4*D))))/(256*D**2*t**4+256*D**2*t**3+96*D**2*t**2+16*D**2*t+D**2)
                # adding noise
                time.append(t*0.1)
                u_noise.append((1+(random.random()-random.random())*0.02*n)*u)               
                u_t_noise.append((1+(random.random()-random.random())*0.02*n)*u_t)
                u_x_noise.append((1+(random.random()-random.random())*0.02*n)*u_x)
                u_y_noise.append((1+(random.random()-random.random())*0.02*n)*u_y) 
                u_z_noise.append((1+(random.random()-random.random())*0.02*n)*u_z)
                u_xx_noise.append((1+(random.random()-2*random.random()+random.random())*0.02*n)*u_xx)
                u_yy_noise.append((1+(random.random()-2*random.random()+random.random())*0.02*n)*u_yy)
                u_zz_noise.append((1+(random.random()-2*random.random()+random.random())*0.02*n)*u_zz)
                
            for k in range(len(time)):
                l = i * 50 + 10 * j + k + 1
                if(n == 0):
                    ws1.cell(l+1,1).value = x
                    ws1.cell(l+1,2).value = y
                    ws1.cell(l+1,3).value = z
                    ws1.cell(l+1,4).value = time[k]
                    ws1.cell(l+1,5).value = u_noise[k]
                    ws1.cell(l+1,6).value = u_t_noise[k]
                    ws1.cell(l+1,7).value = u_x_noise[k]
                    ws1.cell(l+1,8).value = u_y_noise[k]
                    ws1.cell(l+1,9).value = u_z_noise[k]
                    ws1.cell(l+1,10).value = u_xx_noise[k]
                    ws1.cell(l+1,11).value = u_yy_noise[k]
                    ws1.cell(l+1,12).value = u_zz_noise[k]
                elif(n == 10):
                    ws2.cell(l+1,1).value = x
                    ws2.cell(l+1,2).value = y
                    ws2.cell(l+1,3).value = z
                    ws2.cell(l+1,4).value = time[k]
                    ws2.cell(l+1,5).value = u_noise[k]
                    ws2.cell(l+1,6).value = u_t_noise[k]
                    ws2.cell(l+1,7).value = u_x_noise[k]
                    ws2.cell(l+1,8).value = u_y_noise[k]
                    ws2.cell(l+1,9).value = u_z_noise[k]
                    ws2.cell(l+1,10).value = u_xx_noise[k]
                    ws2.cell(l+1,11).value = u_yy_noise[k]
                    ws2.cell(l+1,12).value = u_zz_noise[k]

            time = []
            u_noise = []
            u_t_noise = []
            u_x_noise = []
            u_y_noise = []
            u_z_noise = []
            u_xx_noise = []
            u_yy_noise = []
            u_zz_noise = []

        path1 = "./data/AdvectionDiffusion/noise_00.csv"
        wb1.save(path1)
        path2 = "./data/AdvectionDiffusion/noise_10.csv"
        wb2.save(path2)
