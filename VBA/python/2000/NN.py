"""
----------- Data Generation For Machine Learning ----------
"""
import numpy as np
import math 
import random

# u_t == -Vx * u_x -Vy * u_y -Vz * u_z + D * (u_xx + u_yy + u_zz)
# velocity
Vx = 1
Vy = 2
Vz = 1
# diffusivities
D = 1

# center
center = [0.5, 0.5, 0.5]

# training set
# change for 100 different initial conditions --> center (x0, y0, z0)
for i in range(0,1,1):
    x0, y0, z0 = center[0], center[1], center[2]
    # numbers of dataset in each initial conditions --> position (x, y, z)
    for j in range(0,2000,1):
        x, y, z, t = random.random(), random.random(), random.random(), random.random()
        # exact solution in (x, y, z, t) referring to initial condition (x0, y0, z0)
        u = 1/((4*t+1)**1.5)*math.exp((-(x-Vx*t-x0)**2/(D*(4*t+1))-(y-Vy*t-y0)**2/(D*(4*t+1))-(z-Vz*t-z0)**2/(D*(4*t+1))))
        if j == 0:
            x_tr = np.array([[x, y, z, t]])
            y_tr = np.array([[u]])
        else:
            x_tr = np.vstack((x_tr, np.array([[x, y, z, t]])))
            y_tr = np.vstack((y_tr, np.array([[u]])))

"""
----------- Machine Learning model ----------
"""
import sklearn
import tensorflow as tf
from tensorflow import keras
from sklearn.model_selection import train_test_split
x_train_all, x_test, y_train_all, y_test = train_test_split(x_tr, y_tr, random_state=7)
x_train, x_valid, y_train, y_valid = train_test_split(x_train_all, y_train_all, random_state=11)

# print(x_train.shape, y_train.shape)
# print(x_valid.shape, y_valid.shape)
# print(x_test.shape, y_test.shape)

def build_model(learning_rate = 3e-3):
    model = keras.models.Sequential()
    model.add(keras.layers.Dense(4, activation="sigmoid",
                                input_shape=x_train.shape[1:]))
    model.add(keras.layers.Dense(5,activation = "sigmoid"))
    model.add(keras.layers.Dense(6,activation = "sigmoid"))
    model.add(keras.layers.Dense(5,activation = "sigmoid"))
    model.add(keras.layers.Dense(4,activation = "sigmoid"))
    model.add(keras.layers.Dense(1,activation = "sigmoid"))
    optimizer = keras.optimizers.Adam(learning_rate)
    model.compile(loss = 'mse', optimizer = optimizer)
    return model

callbacks = [keras.callbacks.EarlyStopping(patience=5, min_delta=1e-10)]

model = build_model()
history = model.fit(x_train, y_train, epochs = 100,
                    validation_data = (x_valid, y_valid),
                    callbacks = callbacks)
# print(model.summary())
model.evaluate(x_test, y_test)

from tensorflow import float64
# get weights and bias in each layer
i = 0
kernel_tensor = []
bias_tensor = []
for each in model.weights:
    if i % 2 is 0:
        kernel_tensor.append(tf.constant(each.numpy(), dtype=float64))
    else:
        bias_tensor.append(tf.constant(each.numpy(), dtype=float64))
    i += 1

# reconstruct the machine learning model
def s(input_x, input_y, input_z, input_t):
    i = 0
    for w, b in zip(kernel_tensor, bias_tensor):
        if i == 0:
            input_tensor = input_x*w[0,:] + input_y*w[1,:] + input_z*w[2,:] + input_t*w[3,:] + b
            input_tensor = tf.nn.sigmoid(input_tensor)
        elif i < len(kernel_tensor)-1:
            output_tensor = input_tensor@w + b
            output = tf.nn.sigmoid(output_tensor)
            input_tensor = output
        else:
            output_tensor = input_tensor@w + b
            output = tf.nn.sigmoid(output_tensor)
        i += 1
    return output

"""
----------- Calculate the derivatives ----------
"""
import openpyxl

wb = openpyxl.load_workbook(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx') 
ws = wb['NN']

for i in range(8, 2008, 1):
    input_array = np.array([
        [ws.cell(i, 3).value, ws.cell(i, 4).value, ws.cell(i, 5).value, ws.cell(i, 2).value]
    ])
    input_x = tf.Variable(input_array[:, 0].reshape(1, 1))
    input_y = tf.Variable(input_array[:, 1].reshape(1, 1))
    input_z = tf.Variable(input_array[:, 2].reshape(1, 1))
    input_t = tf.Variable(input_array[:, 3].reshape(1, 1))

    ws.cell(i, 9).value = s(input_x, input_y, input_z, input_t).numpy()[0, 0]

    # first derivatives
    with tf.GradientTape(persistent=True) as tape:
        output = s(input_x, input_y, input_z, input_t)

    ds_x = tape.gradient(output, input_x).numpy()
    ds_y = tape.gradient(output, input_y).numpy()
    ds_z = tape.gradient(output, input_z).numpy()
    ds_t = tape.gradient(output, input_t).numpy()

    del tape
    ws.cell(i, 10).value = ds_x[0, 0]
    ws.cell(i, 11).value = ds_y[0, 0]
    ws.cell(i, 12).value = ds_z[0, 0]
    ws.cell(i, 16).value = ds_t[0, 0]

    # second derivatives
    with tf.GradientTape(persistent=True) as outer_tape:
        with tf.GradientTape(persistent=True) as inner_tape:
            output = s(input_x, input_y, input_z, input_t)
        inner_grads = inner_tape.gradient(output, [input_x, input_y, input_z])
    # outer_grads = [outer_tape.gradient(inner_grad, [input_x, input_y, input_z]) for inner_grad in inner_grads]
    ds_xx = outer_tape.gradient(inner_grads[0], [input_x])[0].numpy()
    ds_yy = outer_tape.gradient(inner_grads[1], [input_y])[0].numpy()
    ds_zz = outer_tape.gradient(inner_grads[2], [input_z])[0].numpy()

    del inner_tape
    del outer_tape
    ws.cell(i, 13).value = ds_xx[0, 0]
    ws.cell(i, 14).value = ds_yy[0, 0]
    ws.cell(i, 15).value = ds_zz[0, 0]

wb.save(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx')