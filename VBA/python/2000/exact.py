import pandas as pd
import numpy as np
import openpyxl
import random

wb = openpyxl.load_workbook(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx') 
ws = wb['Exact']

for i in range(22, 2008, 1):
    t, x, y, z = random.random(), random.random(), random.random(), random.random()
    ws.cell(i, 2).value = t
    ws.cell(i, 3).value = x
    ws.cell(i, 4).value = y
    ws.cell(i, 5).value = z

wb.save(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx')