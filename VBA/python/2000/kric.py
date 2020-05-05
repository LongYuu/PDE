"""
----------- Data Generation For Kriging ----------
"""
import numpy as np
import math 
import random
import openpyxl

# u_t == -Vx * u_x -Vy * u_y -Vz * u_z + D * (u_xx + u_yy + u_zz)
# velocity
Vx = 1
Vy = 2
Vz = 1
# diffusivities
D = 1

wb = openpyxl.load_workbook(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx') 
ws = wb['Kriging  (IC)']

for col in range(8, 2008, 1):
    # center
    center = [ws.cell(col, 6).value, ws.cell(col, 7).value, ws.cell(col, 8).value]

    # training set
    x0, y0, z0 = center[0], center[1], center[2]
    # numbers of dataset in each initial conditions --> position (x, y, z)
    for j in range(0,50,1):
        x, y, z, t = random.random(), random.random(), random.random(), random.random()
        # exact solution in (x, y, z, t) referring to initial condition (x0, y0, z0)
        u = 1/((4*t+1)**1.5)*math.exp((-(x-Vx*t-x0)**2/(D*(4*t+1))-(y-Vy*t-y0)**2/(D*(4*t+1))-(z-Vz*t-z0)**2/(D*(4*t+1))))
        if j == 0:
            x_tr = np.array([[x, y, z, t]])
            y_tr = np.array([[u]])
        else:
            x_tr = np.vstack((x_tr, np.array([[x, y, z, t]])))
            y_tr = np.vstack((y_tr, np.array([[u]])))

    """
    ----------- Kriging model ----------
    """
    import numpy.linalg as la
    from scipy.optimize import leastsq

    def read_data(a,ti):

        data_D= np.zeros([a,3])
        uu = np.zeros(([a,1]))
        for i in range(a):

            data_D[i,:]= x_tr[i*10,[0,1,2]]
            uu[i] = y_tr[i*10+ti]


        n, m = data_D.shape
        D = np.zeros([n, n])
        for i in range(n):
            for j in range(i + 1, n):
                D[i, j] = la.norm(data_D[i, :] - data_D[j, :])
                D[j, i] = D[i, j]

        return uu


    def Kriging(a,x,y,z,t):

        data_D= np.zeros([a,4])
        uu = np.zeros(([a,1]))
        for i in range(a):
                data_D[i,:]= x_tr[i,[0,1,2,3]]
                uu[i] = y_tr[i]


        n, m = data_D.shape
        D = np.zeros([n, n])
        for i in range(n):
            for j in range(i + 1, n):
                D[i, j] = la.norm(data_D[i, :] - data_D[j, :])
                D[j, i] = D[i, j]
        #print(D.shape)

        #distance matrix
        n,m = data_D.shape
        D = np.zeros([n,n])
        for i in range(n):
            for j in range(i+1, n):
                D[i,j] = la.norm(data_D[i,:]-data_D[j,:])
                D[j,i] = D[i,j]

        X =np.asarray([x,y,z])
        D_b = np.zeros([n,1])
        for i in range(n):
            D_b[i] = np.sqrt((x-data_D[i,0])**2 + (y-data_D[i,1])**2 + (z-data_D[i,2])**2 + (t-data_D[i,3])**2)


        #semivariance
        n1, m1 = uu.shape
        Y = np.zeros([n1, n1])
        for i in range(n1):
            for j in range(i + 1, n1):
                Y[i, j] = la.norm(uu[i, :] -uu[j, :])
                Y[j, i] = Y[i, j]
        Y = Y / 2

        #REG
        def v(P, h):
            C_0, C, r = P
            return C_0 + C * (1 - np.exp((-h / r)))

        def error(P, h, y_var):
            return v(P, h) - y_var

        D1 = np.ravel(D)
        Y_ = np.ravel(Y)
        P0 = [0.001, 1, 10]
        para = leastsq(error, P0, args=(D1, Y_))

        C_0, C, r = para[0]

        def Vari(h):
            return C_0 + C * (1 - np.exp((-h / r)))

        # covariance matrix

        M_A = np.linalg.inv(Vari(D))
        #M_b = Vari(D_b)

        M_b = np.zeros([n, 1])
        w = np.zeros([n, 1])
        for i in range(n):
            M_b[i] = C_0 + C * (1 - np.exp((-(np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2 + (t-data_D[i,3])**2)) / r)))
            for j in range(n):
                w[i]= M_A[i,j] * ( C_0 + C * (1 - np.exp((-(np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2 + (t-data_D[i,3])**2)) / r))))

        wx =np.zeros([n,1])
        wy =np.zeros([n,1])
        wz =np.zeros([n,1])
        wxx =np.zeros([n,1])
        wyy =np.zeros([n,1])
        wzz =np.zeros([n,1])
        wt = np.zeros([n,1])

        #w[i] = M_A[i,j] *  C_0 + C * (1 - np.exp((-(np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2)) / r)))

        for i in range(n):
            for j in range(n):
                wx[i] =M_A[i,j] * -C *(x-data_D[i,0])*np.exp(-np.sqrt((x-data_D[i,0])**2 +(y-data_D[i,1])**2+(z-data_D[i,2])**2+(t-data_D[i,3])**2)/r)/(r*np.sqrt((x-data_D[i,0])**2+(y-data_D[i,1])**2+(z-data_D[i,2])**2+(t-data_D[i,3])**2))
                wy[i] = M_A[i,j]* -C * (y - data_D[i, 1]) * np.exp(
                    -np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) / r) / (
                                    r * np.sqrt(
                                (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2))
                wz[i] = M_A[i,j]* -C * (z - data_D[i, 2]) * np.exp(
                    -np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) / r) / (
                                    r * np.sqrt(
                                (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2))

                wxx[i]=M_A[i,j]*C * ((x - data_D[i,0]) ** 2 / ((x - data_D[i,0]) ** 2 + (y - data_D[i,1]) ** 2 + (z - data_D[i,2]) ** 2+(t-data_D[i,3])**2) + r * (x - data_D[i,0]) ** 2 / (
                            (x - data_D[i,0]) ** 2 + (y - data_D[i,1]) ** 2 + (z - data_D[i,2]) ** 2+(t-data_D[i,3])**2) ** (3 / 2) - r / np.sqrt(
                    (x - data_D[i,0]) ** 2 + (y - data_D[i,1]) ** 2 + (z - data_D[i,2]) ** 2+(t-data_D[i,3])**2)) * np.exp(
                    -np.sqrt((x - data_D[i,0]) ** 2 + (y - data_D[i,1]) ** 2 + (z - data_D[i,2]) ** 2+(t-data_D[i,3])**2) / r) / r**2

                wyy[i] =M_A[i,j]*C * ((y - data_D[i, 1]) ** 2 / (
                            (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2 +(t-data_D[i,3])**2) + r * (
                                        y - data_D[i, 1]) ** 2 / (
                                    (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) ** (
                                        3 / 2) - r / np.sqrt(
                    (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2)) * np.exp(
                    -np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) / r) / r ** 2

                wzz[i] =M_A[i,j]* C * ((z - data_D[i, 2]) ** 2 / (
                            (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) + r * (
                                        z - data_D[i, 2]) ** 2 / (
                                    (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) ** (
                                        3 / 2) - r / np.sqrt(
                    (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2)) * np.exp(
                    -np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2+(t-data_D[i,3])**2) / r) / r ** 2

                wt[i] = M_A[i, j] * -C * (t - data_D[i, 3]) * np.exp(
                    -np.sqrt((x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2 + (
                                t - data_D[i, 3]) ** 2) / r) / (
                                r * np.sqrt(
                            (x - data_D[i, 0]) ** 2 + (y - data_D[i, 1]) ** 2 + (z - data_D[i, 2]) ** 2 + (
                                        t - data_D[i, 3]) ** 2))

        #print(wx.shape)

        ux = np.dot(wx.T, uu)
        uy = np.dot(wy.T, uu)
        uz = np.dot(wz.T, uu)

        uxx = np.dot(wxx.T,uu)
        uyy = np.dot(wyy.T,uu)
        uzz = np.dot(wzz.T,uu)

        ut = np.dot(wt.T, uu)

        w = np.dot(M_A, M_b)
        #print(w)

        y_p = np.dot(w.T,uu)

        #print(y_p)

        #print(ux,uy,uz)

        return y_p,ux,uy,uz,uxx,uyy,uzz,ut


    """
    ----------- Calculate the derivatives ----------
    """
    result = Kriging(
        50, 
        ws.cell(col, 3).value, 
        ws.cell(col, 4).value, 
        ws.cell(col, 5).value, 
        ws.cell(col, 2).value)
    
    ws.cell(col, 9).value = result[0][0][0]   # u(x, y, z, t) 

    # first derivatives
    
    ws.cell(col, 10).value = result[1][0][0]    # u_x
    ws.cell(col, 11).value = result[2][0][0]     # u_y
    ws.cell(col, 12).value = result[3][0][0]     # u_z
    ws.cell(col, 16).value = result[7][0][0]     # u_t

    # second derivatives
    
    ws.cell(col, 13).value = result[4][0][0]    # u_xx
    ws.cell(col, 14).value = result[5][0][0]    # u_yy
    ws.cell(col, 15).value = result[6][0][0]    # u_zz

wb.save(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx')