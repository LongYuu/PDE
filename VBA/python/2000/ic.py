import pandas as pd
import numpy as np
import openpyxl
import random

wb = openpyxl.load_workbook(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx') 
ws = wb['NN (IC)']

for i in range(22, 2008, 1):
    x0, y0, z0 = random.random(), random.random(), random.random()
    ws.cell(i, 6).value = x0
    ws.cell(i, 7).value = y0
    ws.cell(i, 8).value = z0

wb.save(r'../../20200502_PDEViz(AdvectionDiffusion).xlsx')